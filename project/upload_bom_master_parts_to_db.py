# -*- coding: utf-8 -*-
"""
Do this to grab all the parts from the BOM Master spreadsheet
and upload them to the database.

@author: Sean McCarthy
"""

import os
import pickle
import platform
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import List, Tuple

import numpy as np
import pandas as pd
import psycopg2
from dotenv import load_dotenv
from O365 import Account
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.worksheet.table import Table
from psycopg2.sql import SQL, Literal
from pyxlsb import open_workbook as open_xlsb

# Insert pythonpath into the front of the PATH environment variable, before importing anything from project/
pythonpath = str(Path(__file__).parent.parent)
try:
    sys.path.index(pythonpath)
except ValueError:
    sys.path.insert(0, pythonpath)

from project.logger_config import logger
from project.utils import (
    Config,
    error_wrapper,
    exit_if_already_running,
)

# Load the .env file
env_file_location = Path(__file__).parent.parent.joinpath(".env")
load_dotenv(env_file_location)

# def insert_path(pythonpath):
#     """Insert pythonpath into the front of the PATH environment
#     variable, before importing anything from project/"""
#     try:
#         sys.path.index(str(pythonpath))
#     except ValueError:
#         sys.path.insert(0, str(pythonpath))


# # For running in development only
# project_folder = Path(__file__).absolute().parent.parent
# ad_hoc_folder = project_folder.joinpath("ad_hoc")
# insert_path(project_folder)

# Whether to delete parts no longer found in the BoM Master spreadsheet, or in work orders, etc.
DELETE_DEPRECATED_PARTS: bool = False

if platform.system() == "Linux":
    ijack_folder = Path(r"/workspace/c_users_sean/IJACK")
else:
    ijack_folder = Path(r"C:\Users\seanm\IJACK")

ijack_production_bom_folder = ijack_folder.joinpath("Production New - a - BOM")
ijack_rcom_folder = ijack_folder.joinpath("RCOM - General")
ijack_rcom_temp_folder = ijack_rcom_folder.joinpath("_temp")
ijack_engineering_parts_images_folder = ijack_folder.joinpath(
    "Engineering CAD - RCOM Part Images/Images from Python"
)

EXCEL_FILENAME = "BOM Master 5.3"
file_in_path_xlsb = ijack_production_bom_folder.joinpath(f"{EXCEL_FILENAME}.xlsb")
file_download_path_xlsb = ijack_rcom_temp_folder.joinpath(f"{EXCEL_FILENAME}.xlsb")
file_out_path_xlsx = ijack_rcom_temp_folder.joinpath(f"Sean XLSX {EXCEL_FILENAME}.xlsx")
file_out_path_pickle = ijack_rcom_temp_folder.joinpath(
    f"Sean Pickled {EXCEL_FILENAME}.pkl"
)
file_out_path_dups_removed = ijack_rcom_temp_folder.joinpath(
    f"Sean {EXCEL_FILENAME} Parts (Duplicates Removed).xlsx"
)

#######################################################################################
# Old columns in database table:
# price_cad (real)
# price_usd (real)
# cad_per_usd (real)
#
# New columns:
# cost to be split between cost_cad and cost_usd
# price_cad and price_usd to be msrp_cad and msrp_usd
# dealer_cost_cad and dealer_cost_usd
# ijack_corp_cost is the transfer price from IJACK Inc to IJACK Corp
# multiplier should be msrp_mult_cad and msrp_mult_usd
# transfer_mult_cad_dealer, transfer_mult_usd_dealer, transfer_mult_inc_to_corp
#######################################################################################

sheets_w_part_nums = {
    "Pricing": {
        "db_table_name": "bom_pricing",
        "part_num_col": "DB",
        "description_col": "DC",
        "msrp_mult_cad_col": "DK",
        "transfer_mult_cad_dealer_col": "DL",
        "msrp_mult_usd_col": "DM",
        "transfer_mult_inc_to_corp_col": "DN",
        "transfer_mult_usd_dealer_col": "DO",
        "cost_cad_col": "DP",
        "msrp_cad_col": "DQ",
        "dealer_cost_cad_col": "DR",
        "cost_usd_col": "DS",
        "msrp_usd_col": "DT",
        "ijack_corp_cost_col": "DU",
        "dealer_cost_usd_col": "DV",
        "is_soft_part": "DW",
        "warehouse_mult": "DX",
        "harmonization_code_col": "DY",
        "country_of_origin_col": "DZ",
        "weight_col": "EB",
        "lead_time_col": "EC",
        "is_usd_col": "ED",
        # Starts with 823E in column D, and ends with 2270M4 in column AK
        # NOTE: the below cols need to be updated EVERY time this is run, since the cols change!!!
        "n_parts_cols": [],
    },
    "Base Powerunit": {
        "db_table_name": "bom_base_powerunit",
        "part_num_col": "DB",
        "description_col": "DC",
        "msrp_mult_cad_col": "DK",
        "transfer_mult_cad_dealer_col": "DL",
        "msrp_mult_usd_col": "DM",
        "transfer_mult_inc_to_corp_col": "DN",
        "transfer_mult_usd_dealer_col": "DO",
        "cost_cad_col": "DP",
        "msrp_cad_col": "DQ",
        "dealer_cost_cad_col": "DR",
        "cost_usd_col": "DS",
        "msrp_usd_col": "DT",
        "ijack_corp_cost_col": "DU",
        "dealer_cost_usd_col": "DV",
        "is_soft_part": "DW",
        "warehouse_mult": "DX",
        "harmonization_code_col": "DY",
        "country_of_origin_col": "DZ",
        "weight_col": "EB",
        "lead_time_col": "EC",
        "is_usd_col": "EJ",
        # NOTE: the below cols need to be updated EVERY time this is run, since the cols change!!!
        "n_parts_cols": list(
            range(column_index_from_string("C") - 1, column_index_from_string("M"))
        ),
    },
    "Powerunit": {
        "db_table_name": "bom_powerunit",
        "part_num_col": "DB",
        "description_col": "DC",
        "msrp_mult_cad_col": "DK",
        "transfer_mult_cad_dealer_col": "DL",
        "msrp_mult_usd_col": "DM",
        "transfer_mult_inc_to_corp_col": "DN",
        "transfer_mult_usd_dealer_col": "DO",
        "cost_cad_col": "DP",
        "msrp_cad_col": "DQ",
        "dealer_cost_cad_col": "DR",
        "cost_usd_col": "DS",
        "msrp_usd_col": "DT",
        "ijack_corp_cost_col": "DU",
        "dealer_cost_usd_col": "DV",
        "is_soft_part": "DW",
        "warehouse_mult": "DX",
        "harmonization_code_col": "DY",
        "country_of_origin_col": "DZ",
        "weight_col": "EB",
        "lead_time_col": "EC",
        "is_usd_col": "EK",
        # NOTE: the below cols need to be updated EVERY time this is run, since the cols change!!!
        "n_parts_cols": list(
            range(column_index_from_string("C") - 1, column_index_from_string("AW"))
        ),
    },
    "Structure": {
        "db_table_name": "bom_structure",
        "part_num_col": "DB",
        "description_col": "DC",
        "msrp_mult_cad_col": "DK",
        "transfer_mult_cad_dealer_col": "DL",
        "msrp_mult_usd_col": "DM",
        "transfer_mult_inc_to_corp_col": "DN",
        "transfer_mult_usd_dealer_col": "DO",
        "cost_cad_col": "DP",
        "msrp_cad_col": "DQ",
        "dealer_cost_cad_col": "DR",
        "cost_usd_col": "DS",
        "msrp_usd_col": "DT",
        "ijack_corp_cost_col": "DU",
        "dealer_cost_usd_col": "DV",
        "is_soft_part": "DW",
        "warehouse_mult": "DX",
        "harmonization_code_col": "DY",
        "country_of_origin_col": "DZ",
        "weight_col": "EB",
        "lead_time_col": "EC",
        "is_usd_col": "EJ",
        # NOTE: the below cols need to be updated EVERY time this is run, since the cols change!!!
        "n_parts_cols": list(
            range(column_index_from_string("C") - 1, column_index_from_string("N"))
        ),
    },
    # Pump Top is different since it has "soft part" rows (e.g. Fluidseal, Garlock, Hallite, Trelleborg)
    "Pump Top": {
        "db_table_name": "bom_pump_top",
        "part_num_col": "DB",
        "description_col": "DC",
        "msrp_mult_cad_col": "DK",
        "transfer_mult_cad_dealer_col": "DL",
        "msrp_mult_usd_col": "DM",
        "transfer_mult_inc_to_corp_col": "DN",
        "transfer_mult_usd_dealer_col": "DO",
        "cost_cad_col": "DP",
        "msrp_cad_col": "DQ",
        "dealer_cost_cad_col": "DR",
        "cost_usd_col": "DS",
        "msrp_usd_col": "DT",
        "ijack_corp_cost_col": "DU",
        "dealer_cost_usd_col": "DV",
        "is_soft_part": "DW",
        "warehouse_mult": "DX",
        "harmonization_code_col": "DY",
        "country_of_origin_col": "DZ",
        "weight_col": "EB",
        "lead_time_col": "EC",
        "is_usd_col": "EJ",
        # NOTE: the below cols need to be updated EVERY time this is run, since the cols change!!!
        "n_parts_cols": list(
            range(column_index_from_string("C") - 1, column_index_from_string("CZ"))
        ),
    },
    # DGAS is different
    "DGAS": {
        "db_table_name": "bom_dgas",
        "part_num_col": "DB",
        "description_col": "DC",
        "msrp_mult_cad_col": "DK",
        "transfer_mult_cad_dealer_col": "DL",
        "msrp_mult_usd_col": "DM",
        "transfer_mult_inc_to_corp_col": "DN",
        "transfer_mult_usd_dealer_col": "DO",
        "cost_cad_col": "DP",
        "msrp_cad_col": "DQ",
        "dealer_cost_cad_col": "DR",
        "cost_usd_col": "DS",
        "msrp_usd_col": "DT",
        "ijack_corp_cost_col": "DU",
        "dealer_cost_usd_col": "DV",
        "is_soft_part": "DW",
        "warehouse_mult": "DX",
        "harmonization_code_col": "DY",
        "country_of_origin_col": "DZ",
        "weight_col": "EB",
        "lead_time_col": "EC",
        "is_usd_col": "EJ",
        # NOTE: the below cols need to be updated EVERY time this is run, since the cols change!!!
        "n_parts_cols": list(
            range(column_index_from_string("C") - 1, column_index_from_string("BI"))
        ),
    },
}

# Database connection settings
HOST_IJ = os.getenv("HOST_IJ")
PORT_IJ = int(os.getenv("PORT_IJ"))
USER_IJ = os.getenv("USER_IJ")
PASS_IJ = os.getenv("PASS_IJ")
DB_IJ = "ijack"


def connect_to_o365(client_id: str, client_secret: str, tenant_id: str) -> Account:
    """
    Establish connection to Microsoft 365 account
    Returns authenticated Account object
    """
    # Initialize the token backend
    # token_backend = FileSystemTokenBackend(token_path=".")

    # Define the required scopes for Excel operations
    # scopes = ["basic", "offline_access", "files.read", "files.read.all"]
    # Need to use the below scopes for SharePoint/OneDrive access
    # For client credentials flow, we need to use .default scope
    scopes = [
        "https://graph.microsoft.com/.default",
        # "https://graph.microsoft.com/User.Read",
        # "https://graph.microsoft.com/Files.Read.All",
    ]

    # Create account object
    account = Account(
        credentials=(client_id, client_secret),
        tenant_id=tenant_id,
        auth_flow_type="credentials",
        # token_backend=token_backend,
        # auth_flow_type="credentials",
        # redirect_uri="https://login.microsoftonline.com/common/oauth2/nativeclient",
    )

    # Authenticate (this will open a web browser for consent)
    if account.authenticate(scopes=scopes):
        return account

    raise Exception("Authentication failed")


def list_onedrive_folders(account: Account) -> None:
    """List all folders in OneDrive"""
    storage = account.storage()
    default_drive = storage.get_default_drive()
    root_folder = default_drive.get_root_folder()
    # Generator
    # child_folders = default_drive.get_child_folders()

    # Recursively list all folders
    def print_folder_structure(folder: Path, indent: str = "") -> None:
        logger.info(f"{indent}📁 {folder.name}")
        for item in folder.get_items():
            if item.is_folder:
                print_folder_structure(item, indent + "  ")
            if item.is_file:
                logger.info(f"{indent}  📄 {item.name}")
                logger.info(item.mime_type)  # print the mime type
            elif item.is_photo:
                logger.info(item.camera_model)  # print some metadata of this photo
            elif item.is_image:
                logger.info(item.dimensions)  # print the image dimensions

    print_folder_structure(root_folder)

    return None


def list_sharepoint_files(
    account: Account, site: str = "ProductionNew", folder_path: str = "General/a - BOM"
):
    """List all files in the BOM folder"""
    try:
        # Get the SharePoint context and site using the full site URL
        sharepoint = account.sharepoint()
        # site = sharepoint.get_site('ijack.sharepoint.com/sites/ProductionNew')
        site = sharepoint.get_site(f"ijack.sharepoint.com:/sites/{site}")

        # Get the document library and navigate to the BOM folder
        doc_lib = site.get_default_document_library()
        if folder_path:
            folder = doc_lib.get_item_by_path(folder_path)
        else:
            folder = doc_lib.get_root_folder()

        logger.info(f"Files in {site} folder '{folder_path}':")
        logger.info("-" * 50)

        # List all items in the folder
        def list_items(folder: Path, indent: int = 0):
            """Recursively list all items in the folder"""
            spaces = " " * indent
            for item in folder.get_items():
                item_name_lower = str(item.name).lower()
                if "images" in item_name_lower or "python" in item_name_lower:
                    logger.info(f"{spaces}Found {item.name}!")
                if item.is_folder:
                    logger.info(f"{spaces}📁 {item.name}/")
                    list_items(item, indent + 2)
                else:
                    logger.info(f"{spaces}📄 {item.name}")

        list_items(folder)

    except Exception as e:
        logger.info(f"Error: {str(e)}")


def download_sharepoint_xlsb_file(
    account: Account,
    filename: str = "BOM Master 5.3.xlsb",
    site: str = "ProductionNew",
    folder_path: str = "General/a - BOM",
    to_path: Path = ijack_rcom_temp_folder,
) -> None:
    """
    Get the BOM Master XLSB file from SharePoint
    """
    try:
        sharepoint = account.sharepoint()
        site = sharepoint.get_site(f"ijack.sharepoint.com:/sites/{site}")
        doc_lib = site.get_default_document_library()

        file_path = f"/{folder_path}/{filename}"
        file_item = doc_lib.get_item_by_path(file_path)

        # Get the content and read into pandas
        # Get the drive item and download content
        # drive_item = file_item.get_drive_item()
        logger.info(f"Downloading file '{file_path}' from SharePoint to '{to_path}'...")
        file_item.download(to_path=to_path)

    except Exception as e:
        logger.info(f"Error accessing file: {str(e)}")

    return None


def download_sharepoint_file_to_path(
    account: Account,
    filename: str = "BOM Master 5.3.xlsb",
    folder: str = "General/a - BOM",
) -> Path:
    """
    Get the BOM Master XLSB file from SharePoint into memory
    Returns:
        BytesIO object containing the file content
    """
    sharepoint = account.sharepoint()
    site = sharepoint.get_site("ijack.sharepoint.com:/sites/ProductionNew")
    doc_lib = site.get_default_document_library()

    file_path = f"/{folder}/{filename}"
    logger.info(f"Downloading file '{file_path}' from SharePoint to memory...")

    file_item = doc_lib.get_item_by_path(file_path)

    # Create a temporary directory and file path
    temp_dir = tempfile.mkdtemp()
    temp_path = Path(temp_dir) / filename

    # Download the file to the temp file
    file_item.download(to_path=temp_dir)

    return temp_path


def get_distinct_parts_and_ids(conn) -> dict:
    """DB query for unique parts and their IDs"""

    sql = """
        select distinct t1.part_num, t1.id
        from public.parts t1
    """
    with conn.cursor() as cursor:
        cursor.execute(sql)
        tuples = cursor.fetchall()

    unique_names_and_ids = {row[0]: row[1] for row in tuples}
    logger.info(
        f"\n\nNumber of parts in the public.parts table: {len(unique_names_and_ids)}"
    )

    return unique_names_and_ids


def get_high_level_part_names_and_ids(conn: psycopg2.extensions.connection) -> dict:
    """DB query for unique high-level part names (not including revisions) and their IDs"""

    sql = """
select t1.part_name, t1.id
from parts t1
inner join (
    --subquery to find the highest part revision for each part name
    select part_name, max(part_rev) as part_rev
    from parts
    group by part_name
) a1
on a1.part_name = t1.part_name
    and a1.part_rev = t1.part_rev
    """
    with conn.cursor() as cursor:
        cursor.execute(sql)
        tuples = cursor.fetchall()

    unique_names_and_ids = {row[0]: row[1] for row in tuples}

    return unique_names_and_ids


def finished_good_name_id_dict(cursor, db_table_name: str) -> dict:
    """DB query for unique finished goods and their IDs"""

    sql = f"""
        select distinct t1.name, t1.id
        from public.{db_table_name} t1
    """
    cursor.execute(sql)
    tuples = cursor.fetchall()

    unique_names_and_ids = {row[0]: row[1] for row in tuples}

    logger.info(
        f"\n\nNumber of finished goods in the table '{db_table_name}': {len(unique_names_and_ids)}"
    )
    return unique_names_and_ids


def get_all_tables_from_workbook(wb: Workbook) -> dict:
    """
    Get all tables from a given workbook. Returns a dictionary of tables.
    """

    # Initialize the dictionary of tables
    tables_dict = {}
    # Go through each worksheet in the workbook
    for ws_name in wb.sheetnames:
        logger.info("")
        logger.info(f"worksheet name: {ws_name}")
        ws = wb[ws_name]
        logger.info(f"tables in worksheet: {len(ws.tables)}")
        # Get each table in the worksheet
        for tbl in ws.tables.values():
            # for table_name, tbl in ws.tables.items():
            # logger.info(f"table name: {table_name}")
            name = tbl.name
            logger.info(f"table name: {name}")
            # First, add some info about the table
            tables_dict[name] = {
                "table_name": name,
                "worksheet": ws_name,
                "num_cols": len(tbl.tableColumns),
                "table_range": tbl.ref,
            }
            # Now convert the table data to a pandas dataframe
            data = ws[tbl.ref]
            # Get a list of all rows, including the first header row
            rows_list = []
            for row in data:
                # Get a list of all columns in each row
                cols = []
                for col in row:
                    cols.append(col.value)
                rows_list.append(cols)
            # Create a pandas dataframe from the rows_list. The first row is the column names
            df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
            # Add the dataframe to the dictionary of tables
            tables_dict[tbl.name]["dataframe"] = df
    logger.info("")
    return tables_dict


def check_for_newline_chars(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove literal newline characters from records.
    Check for "\n" in value or "\r" in value and replace with ' ' blank.
    """
    # Remove literal newline characters from records
    df2 = df.copy()
    for index, row in df.loc[:, :].iterrows():
        # logger.info(f"row index: {index}, row: {row}\n")
        # logger.info(row['abbrev'], row['description'])
        for column, value in row.items():
            if isinstance(value, str) and ("\n" in value or "\r" in value):
                new_string = value.replace("\n", " ").replace("\r", " ")
                df2.loc[index, column] = new_string
                logger.info(f"\n row index: {index}")
                logger.info(f"column: {column}, new_string: {new_string}")

    return df2


def update_parts_table(df: pd.DataFrame, conn) -> None:
    """Update parts from the BOM Master spreadsheet"""

    logger.info("\n\nUpdating parts data in the AWS RDS database...")
    n_rows = len(df)
    assert (
        n_rows > 100
    ), f"Not enough rows in new spreadsheet. Only {n_rows}, so not deleting or updating anything!"

    with conn.cursor() as cursor:
        # Delete all rows not in the new spreadsheet, not in the work orders table,
        # and not marked as "no_delete" in the DB.
        sql_get_unique_records = """
            select distinct on (part_num)
                part_num,
                no_delete,
                id as part_id
            from public.parts
        """
        cursor.execute(sql_get_unique_records)
        tuples = cursor.fetchall()
        unique_parts_db = {row[0]: row[2] for row in tuples}
        # Parts we're not allowed to delete, since they're marked as such in the "no_delete" column
        unique_parts_db_no_delete = [row[0] for row in tuples if row[1] is True]

        sql_get_unique_work_order_parts = """
            select distinct t2.part_num
            from public.work_orders_parts t1
            left join public.parts t2
                on t2.id = t1.part_id
            where t1.part_id is not null
        """
        cursor.execute(sql_get_unique_work_order_parts)
        tuples_wo = cursor.fetchall()
        unique_parts_db_wo = [row[0] for row in tuples_wo]

        sql_get_unique_inventory_parts = """
            select distinct t2.part_num
            from public.warehouses_parts_rel t1
            left join public.parts t2
                on t2.id = t1.part_id
            where t1.part_id is not null
        """
        cursor.execute(sql_get_unique_inventory_parts)
        tuples_inv = cursor.fetchall()
        unique_parts_db_inv = [row[0] for row in tuples_inv]

        unique_parts_new = df["part_num"].drop_duplicates().to_list()
        parts_to_delete = {}
        for part_num, part_id in unique_parts_db.items():
            if all(
                [
                    part_num not in unique_parts_new,
                    part_num not in unique_parts_db_wo,
                    part_num not in unique_parts_db_no_delete,
                    part_num not in unique_parts_db_inv,
                ]
            ):
                parts_to_delete[part_num] = part_id

        n_parts_to_delete = len(parts_to_delete)
        if n_parts_to_delete > 0 and DELETE_DEPRECATED_PARTS is True:
            logger.info(
                f"WARNING: Deleting {n_parts_to_delete} parts which are no longer in the BOM Master spreadsheet!"
            )
            part_id_for_sql = SQL(",").join(map(Literal, parts_to_delete.values()))

            sql_delete = SQL(
                """
                delete from bom_base_powerunit_part_rel where part_id in ({});
                delete from bom_powerunit_part_rel where part_id in ({});
                delete from bom_pricing_part_rel where part_id in ({});
                delete from bom_structure_part_rel where part_id in ({});
                delete from bom_pump_top_part_rel where part_id in ({});
                delete from bom_dgas_part_rel where part_id in ({});
                delete from public.parts where id in ({});
                """
            ).format(
                part_id_for_sql,
                part_id_for_sql,
                part_id_for_sql,
                part_id_for_sql,
                part_id_for_sql,
                part_id_for_sql,
                part_id_for_sql,
            )
            cursor.execute(sql_delete)
            conn.commit()

        counter = 0
        for row in df.itertuples():
            counter += 1
            logger.info(f"Row {counter} of {n_rows}")

            worksheet = str(row.worksheet).replace("'", '"').replace("%", r"%%")
            ws_row = float(row.ws_row)
            part_num = str(row.part_num).replace("'", '"').replace("%", r"%%")
            description = str(row.description).replace("'", '"').replace("%", r"%%")
            msrp_mult_cad = float(row.msrp_mult_cad)
            transfer_mult_cad_dealer = float(row.transfer_mult_cad_dealer)
            msrp_mult_usd = float(row.msrp_mult_usd)
            transfer_mult_inc_to_corp = float(row.transfer_mult_inc_to_corp)
            transfer_mult_usd_dealer = float(row.transfer_mult_usd_dealer)
            warehouse_mult = float(row.warehouse_mult)
            cost_cad = float(row.cost_cad)
            msrp_cad = float(row.msrp_cad)
            dealer_cost_cad = float(row.dealer_cost_cad)
            cost_usd = float(row.cost_usd)
            msrp_usd = float(row.msrp_usd)
            ijack_corp_cost = float(row.ijack_corp_cost)
            dealer_cost_usd = float(row.dealer_cost_usd)
            is_usd = bool(row.is_usd)
            cad_per_usd = float(row.cad_per_usd)
            is_soft_part = bool(row.is_soft_part)
            weight = row.weight  # some weights are string comments like "per foot"
            lead_time = float(row.lead_time)
            harmonization_code = (
                str(row.harmonization_code).replace("'", '"').replace("%", r"%%")
            )
            country_of_origin = (
                str(row.country_of_origin).replace("'", '"').replace("%", r"%%")
            )

            values = {
                "worksheet": worksheet,
                "ws_row": ws_row,
                "part_num": part_num,
                "description": description,
                "msrp_mult_cad": msrp_mult_cad,
                "transfer_mult_cad_dealer": transfer_mult_cad_dealer,
                "msrp_mult_usd": msrp_mult_usd,
                "transfer_mult_inc_to_corp": transfer_mult_inc_to_corp,
                "transfer_mult_usd_dealer": transfer_mult_usd_dealer,
                "warehouse_mult": warehouse_mult,
                "cost_cad": cost_cad,
                "msrp_cad": msrp_cad,
                "dealer_cost_cad": dealer_cost_cad,
                "cost_usd": cost_usd,
                "msrp_usd": msrp_usd,
                "ijack_corp_cost": ijack_corp_cost,
                "dealer_cost_usd": dealer_cost_usd,
                "is_usd": is_usd,
                "cad_per_usd": cad_per_usd,
                "is_soft_part": is_soft_part,
                "harmonization_code": harmonization_code,
                "country_of_origin": country_of_origin,
                "weight": weight,
                "lead_time": lead_time,
            }

            sql_update_existing_parts = """
                INSERT INTO public.parts
                    (worksheet, ws_row, part_num, description, msrp_mult_cad, transfer_mult_cad_dealer, msrp_mult_usd, transfer_mult_inc_to_corp, transfer_mult_usd_dealer, warehouse_mult, cost_cad, msrp_cad, dealer_cost_cad, cost_usd, msrp_usd, ijack_corp_cost, dealer_cost_usd, is_usd, cad_per_usd, is_soft_part, weight, harmonization_code, country_of_origin, lead_time)
                VALUES
                    (%(worksheet)s, %(ws_row)s, %(part_num)s, %(description)s, %(msrp_mult_cad)s, %(transfer_mult_cad_dealer)s, %(msrp_mult_usd)s, %(transfer_mult_inc_to_corp)s, %(transfer_mult_usd_dealer)s, %(warehouse_mult)s, %(cost_cad)s, %(msrp_cad)s, %(dealer_cost_cad)s, %(cost_usd)s, %(msrp_usd)s, %(ijack_corp_cost)s, %(dealer_cost_usd)s, %(is_usd)s, %(cad_per_usd)s, %(is_soft_part)s, %(weight)s, %(harmonization_code)s, %(country_of_origin)s, %(lead_time)s)
                ON CONFLICT (part_num) DO UPDATE
                    SET
                        worksheet = %(worksheet)s,
                        ws_row = %(ws_row)s,
                        description = %(description)s,
                        msrp_mult_cad = %(msrp_mult_cad)s,
                        transfer_mult_cad_dealer = %(transfer_mult_cad_dealer)s,
                        msrp_mult_usd = %(msrp_mult_usd)s,
                        transfer_mult_inc_to_corp = %(transfer_mult_inc_to_corp)s,
                        transfer_mult_usd_dealer = %(transfer_mult_usd_dealer)s,
                        warehouse_mult = %(warehouse_mult)s,
                        cost_cad = %(cost_cad)s,
                        msrp_cad = %(msrp_cad)s,
                        dealer_cost_cad = %(dealer_cost_cad)s,
                        cost_usd = %(cost_usd)s,
                        msrp_usd = %(msrp_usd)s,
                        ijack_corp_cost = %(ijack_corp_cost)s,
                        dealer_cost_usd = %(dealer_cost_usd)s,
                        is_usd = %(is_usd)s,
                        cad_per_usd = %(cad_per_usd)s,
                        is_soft_part = %(is_soft_part)s,
                        weight = %(weight)s,
                        lead_time = %(lead_time)s,
                        harmonization_code = %(harmonization_code)s,
                        country_of_origin = %(country_of_origin)s
            """
            cursor.execute(sql_update_existing_parts, values)
        conn.commit()

    return None


def get_all_tables_from_dict(tables_dict: dict) -> pd.DataFrame:
    """Get all tables from Excel file"""

    parts_df: pd.DataFrame = tables_dict["Table1"]["dataframe"]
    parts_df.columns = [x.lower().replace(" ", "_") for x in parts_df.columns]
    parts_cols = [
        "worksheet",
        "ws_row",
        "part_num",
        "description",
        "msrp_mult_cad",
        "transfer_mult_cad_dealer",
        "msrp_mult_usd",
        "transfer_mult_inc_to_corp",
        "transfer_mult_usd_dealer",
        "warehouse_mult",
        "cost_cad",
        "msrp_cad",
        "dealer_cost_cad",
        "cost_usd",
        "msrp_usd",
        "ijack_corp_cost",
        "dealer_cost_usd",
        "is_usd",
        "cad_per_usd",
        "is_soft_part",
        "harmonization_code",
        "country_of_origin",
        "weight",
        "lead_time",
    ]
    parts_df = parts_df[parts_cols]
    parts_df.info()

    return parts_df


def add_unique_finished_good(db_table_name: str, unique_name: str, cursor) -> int:
    """Upload a unique finished good to the database and return its ID"""

    logger.info(
        f"Inserting unique finished good '{unique_name}' to table '{db_table_name}'..."
    )
    sql_insert = f"""
        INSERT INTO public.{db_table_name}
            (name)
        VALUES
            (%(name)s)
        ON CONFLICT (name) DO NOTHING
        RETURNING id
    """
    values = {"name": unique_name}
    cursor.execute(sql_insert, values)
    # conn.commit()
    id = cursor.fetchone()[0]

    return id


def go_through_all_sheets(
    sheets_w_part_nums: dict, wb: Workbook, cad_per_usd: float
) -> Tuple[List[dict], dict]:
    """Go through each worksheet in the workbook"""

    part_list_of_dicts: list = []
    # This is going to get uploaded to the relational tables, BY WORKSHEET
    finished_goods_dict: dict = {}

    for ws_name, ws_dict in sheets_w_part_nums.items():
        ws = wb[ws_name]
        logger.info(f"\n\n--------- Worksheet name: {ws_name} ---------")
        worksheet_parts_counter = 0
        # This list is just for counting UNIQUE parts in each worksheet,
        # for comparison with the above simple parts counter. It does NOT get uploaded anywhere.
        worksheet_list_of_dicts = []

        db_table_name: str = ws_dict["db_table_name"]
        finished_goods_dict[db_table_name] = []
        n_parts_cols_range = ws_dict["n_parts_cols"]

        # Column numbers, from letters
        # logger.info(f"Testing: column 'a' is column number: {column_index_from_string('a')}")
        # logger.info(f"Testing: column 'B' is column number: {column_index_from_string('B')}")
        part_num_col_str = ws_dict["part_num_col"]
        part_num_col = column_index_from_string(part_num_col_str)

        description_col_str = ws_dict["description_col"]
        description_col = column_index_from_string(description_col_str)

        msrp_mult_cad_col_str = ws_dict["msrp_mult_cad_col"]
        msrp_mult_cad_col = column_index_from_string(msrp_mult_cad_col_str)

        transfer_mult_cad_dealer_col_str = ws_dict["transfer_mult_cad_dealer_col"]
        transfer_mult_cad_dealer_col = column_index_from_string(
            transfer_mult_cad_dealer_col_str
        )

        msrp_mult_usd_col_str = ws_dict["msrp_mult_usd_col"]
        msrp_mult_usd_col = column_index_from_string(msrp_mult_usd_col_str)

        transfer_mult_inc_to_corp_col_str = ws_dict["transfer_mult_inc_to_corp_col"]
        transfer_mult_inc_to_corp_col = column_index_from_string(
            transfer_mult_inc_to_corp_col_str
        )

        transfer_mult_usd_dealer_col_str = ws_dict["transfer_mult_usd_dealer_col"]
        transfer_mult_usd_dealer_col = column_index_from_string(
            transfer_mult_usd_dealer_col_str
        )

        is_soft_part_col_str = ws_dict["is_soft_part"]
        is_soft_part_col = column_index_from_string(is_soft_part_col_str)

        warehouse_mult_col_str = ws_dict["warehouse_mult"]
        warehouse_mult_col = column_index_from_string(warehouse_mult_col_str)

        cost_cad_col_str = ws_dict["cost_cad_col"]
        cost_cad_col = column_index_from_string(cost_cad_col_str)

        msrp_cad_col_str = ws_dict["msrp_cad_col"]
        msrp_cad_col = column_index_from_string(msrp_cad_col_str)

        dealer_cost_cad_col_str = ws_dict["dealer_cost_cad_col"]
        dealer_cost_cad_col = column_index_from_string(dealer_cost_cad_col_str)

        cost_usd_col_str = ws_dict["cost_usd_col"]
        cost_usd_col = column_index_from_string(cost_usd_col_str)

        msrp_usd_col_str = ws_dict["msrp_usd_col"]
        msrp_usd_col = column_index_from_string(msrp_usd_col_str)

        ijack_corp_cost_col_str = ws_dict["ijack_corp_cost_col"]
        ijack_corp_cost_col = column_index_from_string(ijack_corp_cost_col_str)

        dealer_cost_usd_col_str = ws_dict["dealer_cost_usd_col"]
        dealer_cost_usd_col = column_index_from_string(dealer_cost_usd_col_str)

        is_usd_col_str = ws_dict["is_usd_col"]
        is_usd_col = column_index_from_string(is_usd_col_str)

        weight_col_str = ws_dict["weight_col"]
        weight_col = column_index_from_string(weight_col_str)

        lead_time_col_str = ws_dict["lead_time_col"]
        lead_time_col = column_index_from_string(lead_time_col_str)

        harmonization_code_col_str = ws_dict["harmonization_code_col"]
        harmonization_code_col = column_index_from_string(harmonization_code_col_str)

        country_of_origin_col_str = ws_dict["country_of_origin_col"]
        country_of_origin_col = column_index_from_string(country_of_origin_col_str)

        # For each row in the worksheet
        for row in ws.iter_rows(
            min_row=3,
            # min_col=is_usd_col,
            min_col=part_num_col,
            max_col=ws.max_column,
            # max_col=is_usd_col + 1,
            max_row=ws.max_row,
            # max_row=17,
            values_only=False,
        ):
            # Go through each column in the row from left to right, and break once we've added the part from that row
            for cell in row[1:]:
                if cell.value is None:
                    continue  # move on to the next column in the row

                # Find the part_num and check if it's empty
                row_num = cell.row
                part_num = ws.cell(row_num, part_num_col).value
                if part_num is None:
                    continue
                part_num = str(part_num).strip()
                description = ws.cell(row_num, description_col).value
                msrp_mult_cad = ws.cell(row_num, msrp_mult_cad_col).value
                transfer_mult_cad_dealer = ws.cell(
                    row_num, transfer_mult_cad_dealer_col
                ).value
                msrp_mult_usd = ws.cell(row_num, msrp_mult_usd_col).value
                transfer_mult_inc_to_corp = ws.cell(
                    row_num, transfer_mult_inc_to_corp_col
                ).value
                transfer_mult_usd_dealer = ws.cell(
                    row_num, transfer_mult_usd_dealer_col
                ).value

                # Is this a soft part?
                is_soft_part: bool = False
                # if ws_name == "Pump Top":
                #     is_soft_part = row_num in pump_top_all_soft_part_rows
                # elif ws_name == "DGAS":
                #     is_soft_part = row_num in dgas_all_soft_part_rows

                is_soft_part = ws.cell(row_num, is_soft_part_col).value
                if isinstance(is_soft_part, str) or is_soft_part is None:
                    is_soft_part = False

                warehouse_mult = ws.cell(row_num, warehouse_mult_col).value
                # The Pricing sheet doesn't have a warehouse_mult column yet
                if (
                    isinstance(warehouse_mult, str)
                    or warehouse_mult is None
                    or np.isnan(warehouse_mult)
                ):
                    warehouse_mult = 0.0

                cost_cad = ws.cell(row_num, cost_cad_col).value
                msrp_cad = ws.cell(row_num, msrp_cad_col).value
                dealer_cost_cad = ws.cell(row_num, dealer_cost_cad_col).value
                cost_usd = ws.cell(row_num, cost_usd_col).value
                msrp_usd = ws.cell(row_num, msrp_usd_col).value
                ijack_corp_cost = ws.cell(row_num, ijack_corp_cost_col).value
                dealer_cost_usd = ws.cell(row_num, dealer_cost_usd_col).value
                # Is the row in USD?
                is_usd = ws.cell(row_num, is_usd_col).value == "USD"
                weight = ws.cell(row_num, weight_col).value
                lead_time = ws.cell(row_num, lead_time_col).value
                harmonization_code = ws.cell(row_num, harmonization_code_col).value
                country_of_origin = ws.cell(row_num, country_of_origin_col).value

                if part_num and cost_cad is not None and msrp_mult_cad is not None:
                    logger.info(f"\n{cell}")
                    logger.info(f"cad_per_usd: {cad_per_usd}")
                    logger.info(f"type(cad_per_usd): {type(cad_per_usd)}")
                    try:
                        cost_cad = float(cost_cad)
                        msrp_mult_usd = float(msrp_mult_usd)
                    except Exception as err:
                        logger.info(
                            f"Error converting cost_cad or msrp_mult_usd to float: {err}"
                        )
                        logger.info(f"cost_cad: {cost_cad}")
                        logger.info(f"msrp_mult_usd: {msrp_mult_usd}")
                        continue

                    # Store the values from this row in a dictionary
                    d = {
                        "worksheet": ws_name,
                        "ws_row": row_num,
                        "part_num": str(part_num).strip(),
                        "description": description,
                        "msrp_mult_cad": float(msrp_mult_cad),
                        "transfer_mult_cad_dealer": float(transfer_mult_cad_dealer),
                        "msrp_mult_usd": float(msrp_mult_usd),
                        "transfer_mult_inc_to_corp": float(transfer_mult_inc_to_corp),
                        "transfer_mult_usd_dealer": float(transfer_mult_usd_dealer),
                        "warehouse_mult": float(warehouse_mult),
                        "cost_cad": float(cost_cad),
                        "msrp_cad": float(msrp_cad),
                        "dealer_cost_cad": float(dealer_cost_cad),
                        "cost_usd": float(cost_usd),
                        "msrp_usd": float(msrp_usd),
                        "ijack_corp_cost": float(ijack_corp_cost),
                        "dealer_cost_usd": float(dealer_cost_usd),
                        "is_usd": is_usd,
                        "cad_per_usd": cad_per_usd,
                        "is_soft_part": is_soft_part,
                        "weight": weight,
                        "lead_time": float(lead_time),
                        "harmonization_code": harmonization_code,
                        "country_of_origin": country_of_origin,
                    }
                    part_list_of_dicts.append(d)
                    worksheet_list_of_dicts.append(d)
                    worksheet_parts_counter += 1
                    # finished_goods_counter: int = 0

                    # Add the finished good unique name and ID
                    # so we can record the number of each part per unique finished good.
                    for col_num in n_parts_cols_range:
                        # Get the part number for this column
                        unique_name = ws.cell(1, col_num).value
                        if unique_name is None:
                            # No part number in row 1 of this column, so move on to the next column
                            continue
                        unique_name = str(unique_name).strip()
                        if len(unique_name) == 0 or unique_name in (" ", "blank"):
                            continue
                        # Get the quantity for this part row and column
                        quantity = ws.cell(row_num, col_num).value
                        # logger.info(f"quantity: {quantity}")
                        if quantity is None:
                            # Don't add this part if the quantity is blank
                            continue
                        try:
                            quantity = float(quantity)
                        except Exception as err:
                            logger.info(
                                f"Error converting quantity '{quantity}' to float: {err}. Skipping..."
                            )
                            continue
                        if quantity == 0.0:
                            # Don't add this part if the quantity is zero
                            continue
                        # Store the values from this row in a dictionary for relational many-to-many table upload
                        d2 = dict(
                            finished_good_name=unique_name,
                            part_num=str(part_num).strip(),
                            quantity=quantity,
                        )
                        finished_goods_dict[db_table_name].append(d2)

                # break once we've added the part from that row (i.e. stop processing this row and move on to next in the outer loop)
                break

        logger.info(f"\n\nFound {worksheet_parts_counter} parts in worksheet {ws_name}")
        logger.info(
            f"Found {len(set([d['part_num'] for d in worksheet_list_of_dicts]))} unique part numbers in worksheet {ws_name}"
        )
        logger.info(f"Found {len(part_list_of_dicts)} parts in total so far")
        logger.info(
            f"Found {len(set([d['part_num'] for d in part_list_of_dicts]))} unique part numbers so far"
        )
        logger.info("-------------------------------------------------------------\n\n")

        ###############################################################################################################
        # Now you have to open the Excel file "file_out_str", highlight the data,
        # and press CTRL+T to make it a table named "Table1" #
        # Then go to "C:\Users\SeanMcCarthy\git\upload-spreadsheets\upload\main.py" to upload the data.
        # Then go to "C:\Users\SeanMcCarthy\IJACK\Service - Documents\Work Orders\IJACK Work Order Form.xlsm" and refresh the table data.
        ###############################################################################################################

        # msg = f"All done! Now you have to open the Excel file '{file_out_path_dups_removed}', "
        # msg += "highlight the data, and press CTRL+T to make it a table named 'Table1'..."
        # logger.info(msg)

    return part_list_of_dicts, finished_goods_dict


def send_to_excel_and_add_tables(
    df: pd.DataFrame, file_out_path_dups_removed: Path
) -> Workbook:
    """Send to Excel after making the table (NOTE: why?)"""

    # df.to_excel(file_out_path_dups_removed, index=False, freeze_panes=(1, 3))
    with pd.ExcelWriter(file_out_path_dups_removed, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, freeze_panes=(1, 3))

        # Get the xlsxwriter workbook and worksheet objects.
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        # Get the dimensions of the dataframe.
        (max_row, max_col) = df.shape
        max_col_letter = get_column_letter(max_col)

        # Create a list of column headers, to use in add_table().
        column_settings = []
        for header in df.columns:
            column_settings.append({"header": header})

        table_range = f"A1:{max_col_letter}{max_row + 1}"
        tab = Table(displayName="Table1", ref=table_range)

        # # Add a default style with striped rows and banded columns
        # style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
        #                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        # tab.tableStyleInfo = style

        """
        Table must be added using ws.add_table() method to avoid duplicate names.
        Using this method ensures table name is unque through out defined names and all other table name.
        """
        worksheet.add_table(tab)

        # # Add the table.
        # worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

        # # Make the columns wider for clarity.
        # worksheet.set_column(0, max_col - 1, 12)

        # Close the Pandas Excel writer and output the Excel file.
        writer.save()

    return workbook


def upsert_finished_good_pairs(
    finished_goods_dict: list,
    part_id_dict: dict,
    conn: psycopg2.extensions.connection,
) -> None:
    """
    Now that the parts are uploaded, get their part IDs and
    upload the part_num/pump_top_id pairs for preventative maintenance
    """

    with conn.cursor() as cursor:
        for db_table_name, finished_goods_list in finished_goods_dict.items():
            n_items_to_insert = len(finished_goods_list)
            many_to_many_table_name: str = f"{db_table_name}_part_rel"
            if n_items_to_insert == 0:
                logger.info(
                    f"No finished goods pairs to upsert into table '{many_to_many_table_name}'. Skipping..."
                )
                continue
            finished_name_id_dict: dict = finished_good_name_id_dict(
                cursor, db_table_name
            )
            logger.info(
                f"Inserting {n_items_to_insert} many-to-many pairs with part quantities, into table '{many_to_many_table_name}'..."
            )
            counter: int = 0
            for dict_ in finished_goods_list:
                counter += 1

                finished_good_name: str = dict_["finished_good_name"]
                part_num: str = dict_["part_num"]
                quantity: float = dict_["quantity"]
                logger.info(
                    f"{counter} of {n_items_to_insert} for {many_to_many_table_name}: {dict_}"
                )

                # Get finished good ID from dictionary
                finished_good_id: int = finished_name_id_dict.get(
                    finished_good_name, None
                )
                if finished_good_id is None:
                    logger.info(
                        f"Finished good ID for finished_good_name '{finished_good_name}' not found in '{db_table_name}'. Uploading now..."
                    )
                    finished_good_id = add_unique_finished_good(
                        db_table_name, finished_good_name, cursor=cursor
                    )
                    # Add the new finished good ID to the dictionary so we don't try to add it again
                    finished_name_id_dict[finished_good_name] = finished_good_id
                if finished_good_id is None or not isinstance(finished_good_id, int):
                    raise TypeError(
                        f"Could not find or add unique finished good '{finished_good_name}' to table '{db_table_name}'!"
                    )

                part_id: int = part_id_dict.get(part_num, None)
                if part_id is None or not isinstance(part_id, int):
                    raise TypeError(
                        f"Could not find part ID for part_num '{part_num}' in the public.parts table!"
                    )

                values = {
                    "finished_good_id": finished_good_id,
                    "part_id": part_id,
                    "quantity": quantity,
                }

                sql_insert = f"""
                    INSERT INTO public.{many_to_many_table_name}
                        (finished_good_id, part_id, quantity)
                    VALUES
                        (%(finished_good_id)s, %(part_id)s, %(quantity)s)
                    ON CONFLICT (finished_good_id, part_id) DO UPDATE
                        SET
                            quantity = {quantity}
                """

                cursor.execute(sql_insert, values)
            conn.commit()

    return None


def upload_part_pictures(
    account: Account,
    conn: psycopg2.extensions.connection,
) -> None:
    """
    Now that the parts are uploaded, upload the part pictures
    from a folder that Tyler Kenny created.
    Search the ijack_engineering_parts_images_folder for all images
    and upload them to the parts table in the database
    with the part_id as the key.
    """
    # For each part_name, get the part_id with the most recent revision
    part_name_id_dict: dict = get_high_level_part_names_and_ids(conn)

    pictures_without_part_id = {}

    # list_sharepoint_files(account=account, site="EngineeringCAD", folder_path="")

    for image_path in ijack_engineering_parts_images_folder.glob("*.png"):
        # Get the part number from the image name
        part_num = image_path.stem
        part_id = part_name_id_dict.get(part_num, None)
        if part_id is None:
            logger.info(
                f"Part number '{part_num}' not found in the public.parts table!"
            )
            pictures_without_part_id[part_num] = image_path
            continue
        with open(image_path, "rb") as image:
            image_data = image.read()
        with conn.cursor() as cursor:
            cursor.execute(
                """
                UPDATE public.parts
                SET part_image = %s
                WHERE id = %s
                """,
                (image_data, part_id),
            )
        conn.commit()

    if len(pictures_without_part_id) > 0:
        list_str = SQL(",").join(map(Literal, pictures_without_part_id.keys()))
        sql = SQL("""
        SELECT id as part_id, part_name, part_num
        FROM public.parts
        WHERE part_name IN ({list_str})""").format(list_str=list_str)
        # Convert the SQL object to a string using the connection
        query_string = sql.as_string(conn)
        logger.info(f"Query for images without matching parts: {query_string}")
        with conn.cursor() as cursor:
            cursor.execute(sql)
            tuples = cursor.fetchall()
        logger.info(f"Search results for images without matching parts: {tuples}")

    return None


def get_workbook_from_files(
    file_in_path_xlsb: Path, file_out_path_xlsx: Path, file_out_path_pickle: Path
) -> Workbook:
    """Open and write file XLSX file"""

    if file_out_path_pickle.is_file():
        logger.info(
            f"Loading XLSX file with pickle from '{file_out_path_pickle}'. Takes awhile..."
        )
        wb: Workbook = pickle.load(open(file_out_path_pickle, "rb"))
        return wb

    # if False:
    if not file_out_path_xlsx.exists():
        sheets = {}
        logger.info("Extracting data from XLSB file. Doesn't take very long...")
        with open_xlsb(file_in_path_xlsb) as wb:
            for name in wb.sheets:
                sheets[name] = {}
                with wb.get_sheet(name) as sheet:
                    rows = []
                    for row in sheet.rows():
                        rows.append([item.v for item in row])
                    sheets[name]["cols"] = rows[0]
                    sheets[name]["rows"] = rows[1:]

        logger.info("Writing to XLSX file. Takes awhile...")
        with pd.ExcelWriter(file_out_path_xlsx, engine="openpyxl") as writer:
            for ws_name, dict_ in sheets.items():
                df = pd.DataFrame(data=dict_["rows"], columns=dict_["cols"])
                if df.empty:
                    logger.info(f"Empty DataFrame for sheet '{ws_name}'. Skipping...")
                    continue

                # Drop first row of dataframe
                df = df.iloc[:, 1:]

                df.to_excel(writer, sheet_name=ws_name)

    logger.info(
        f"Loading XLSX file from '{file_out_path_xlsx}' or '{file_out_path_pickle}'. Takes awhile..."
    )
    # content_bytes = file_in_path_xlsb.read_bytes()
    # io_bytes = io.BytesIO(content_bytes)
    # wb: Workbook = load_workbook(io_bytes, read_only=False, data_only=True)
    # wb: Workbook = pd.read_excel(file_in_path_xlsb, engine="pyxlsb")
    wb: Workbook = load_workbook(
        filename=file_out_path_xlsx,
        # filename=xlsx_data,
        # Consider changing read_only to False for better performance (counter-intuitive...)
        read_only=False,
        keep_vba=False,
        data_only=True,
        keep_links=False,
    )

    # logger.info(f"Pickling XLSX file (faster load times) to '{file_out_path_pickle}'...")
    # pickle.dump(wb, open(file_out_path_pickle, "wb"))

    if file_out_path_xlsx.is_file():
        logger.info(
            "\nRemoving unnecessary XLSX file so no-one is confused which one to use in the future..."
        )
        file_out_path_xlsx.unlink()

    return wb


def get_workbook_from_xlsb(file_in_path_xlsb: Path) -> Workbook:
    """Convert XLSB file to Workbook object entirely in memory"""
    logger.info("Extracting data from XLSB file...")

    # First read XLSB into dictionary of sheets
    sheets = {}
    with open_xlsb(file_in_path_xlsb) as wb:
        for name in wb.sheets:
            sheets[name] = {}
            with wb.get_sheet(name) as sheet:
                rows = []
                for row in sheet.rows():
                    rows.append([item.v for item in row])
                sheets[name]["cols"] = rows[0]
                sheets[name]["rows"] = rows[1:]

    # Create new workbook in memory
    new_wb = Workbook()

    # Remove default sheet
    new_wb.remove(new_wb.active)

    # Create sheets and populate data
    for ws_name, dict_ in sheets.items():
        # Convert to DataFrame first (easier to handle)
        df = pd.DataFrame(data=dict_["rows"], columns=dict_["cols"])

        if df.empty:
            logger.info(f"Empty DataFrame for sheet '{ws_name}'. Skipping...")
            continue

        # Drop first row of dataframe as per original code
        # df = df.iloc[:, 1:]

        # Create new sheet
        ws = new_wb.create_sheet(title=ws_name)

        # Write headers to first row
        for col_idx, column in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=column)

        # Write data starting from second row
        for row_idx, row in enumerate(df.values, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    return new_wb


def entrypoint(
    # c: Config = None
):
    """Main function"""

    # Assign the Config object to the global variable
    # c = c or Config()

    start_time = datetime.now()
    with psycopg2.connect(
        host=HOST_IJ,
        port=int(PORT_IJ),
        dbname=DB_IJ,
        user=USER_IJ,
        password=PASS_IJ,
        connect_timeout=10,
    ) as conn:
        # Just print how many parts are in the database before we start
        # We'll re-create this dictionary after new parts have been upserted
        part_id_dict: dict = get_distinct_parts_and_ids(conn=conn)

        # Connect to the Office 365 account and download the BOM Master file
        account: Account = connect_to_o365(
            client_id=os.getenv("AZURE_BOM_MASTER_APP_ID"),
            client_secret=os.getenv("AZURE_BOM_MASTER_CLIENT_SECRET"),
            tenant_id=os.getenv("AZURE_BOM_MASTER_TENANT_ID"),
        )

        # Get the BOM Master file from SharePoint
        # download_sharepoint_xlsb_file(
        #     account, filename="BOM Master 5.3.xlsb", to_path=ijack_rcom_temp_folder
        # )
        temp_file_path: Path = download_sharepoint_file_to_path(
            account, filename="BOM Master 5.3.xlsb"
        )

        # # First convert the XLSB file to XLSX so OpenPyXL can read it
        # wb: Workbook = get_workbook_from_files(
        #     # file_in_path_xlsb=file_in_path_xlsb,
        #     file_in_path_xlsb=file_download_path_xlsb,
        #     file_out_path_xlsx=file_out_path_xlsx,
        #     file_out_path_pickle=file_out_path_pickle,
        # )

        wb: Workbook = get_workbook_from_xlsb(file_in_path_xlsb=temp_file_path)

        # Check USD FX rate
        cad_per_usd = wb["Setup"]["b2"].value
        logger.info(f"cad_per_usd exchange rate: {cad_per_usd}")
        assert isinstance(cad_per_usd, float)
        assert 1.0 < cad_per_usd < 1.7

        part_list_of_dicts: List[dict]
        finished_goods_dict: dict
        part_list_of_dicts, finished_goods_dict = go_through_all_sheets(
            sheets_w_part_nums=sheets_w_part_nums, wb=wb, cad_per_usd=cad_per_usd
        )

        # Make a new Pandas DataFrame to store values in memory
        # Sort by part number and cost_cad, descending so highest-cost_cad first
        parts_df = pd.DataFrame(part_list_of_dicts).sort_values(
            ["part_num", "cost_cad"], ascending=False
        )
        # Keep only the highest-cost_cad of each part number
        parts_df = parts_df.drop_duplicates(subset="part_num", keep="first")
        logger.info(
            f"Found {len(parts_df)} unique part numbers after removing duplicates"
        )

        parts_df_no_newline: pd.DataFrame = check_for_newline_chars(parts_df)

        # if True:
        #     # This stuff doesn't take long, but it's unnecessary if you're just uploading the parts
        #     workbook_2: Workbook = send_to_excel_and_add_tables(
        #         df=parts_df, file_out_path_dups_removed=file_out_path_dups_removed
        #     )

        #     tables_dict: dict = get_all_tables_from_workbook(wb=workbook_2)

        #     parts_table_df: pd.DataFrame = get_all_tables_from_dict(tables_dict=tables_dict)

        #     parts_df_no_newline: pd.DataFrame = check_for_newline_chars(parts_table_df)

        # Upload all parts to database 'parts' table.
        update_parts_table(df=parts_df_no_newline, conn=conn)

        # Re-create the part_id_dict after new parts have been upserted
        part_id_dict: dict = get_distinct_parts_and_ids(conn=conn)

        # Now that the parts are uploaded, get their part IDs and upload the part_num/pump_top_id pairs
        # for preventative maintenance
        logger.info(
            "\n\nUpserting finished goods relation tables with part quantities..."
        )
        upsert_finished_good_pairs(
            finished_goods_dict=finished_goods_dict,
            part_id_dict=part_id_dict,
            conn=conn,
        )

        # Now that the parts are uploaded, upload the part pictures
        # upload_part_pictures(conn=conn, account=account)

    if file_out_path_dups_removed.is_file():
        logger.info(
            "\nRemoving unnecessary duplicates-removed Excel file since everything's already been uploaded..."
        )
        file_out_path_dups_removed.unlink()

    if file_out_path_xlsx.is_file():
        logger.info(
            "\nRemoving unnecessary XLSX file so no-one is confused which one to use in the future..."
        )
        file_out_path_xlsx.unlink()

    if file_out_path_pickle.is_file():
        logger.info("\nRemoving unnecessary pickle file...")
        file_out_path_pickle.unlink()

    if temp_file_path.is_file():
        logger.info("\nRemoving unnecessary temp XLSB file...")
        temp_file_path.unlink()

    end_time = datetime.now()
    mins_elapsed = round((end_time - start_time).seconds / 60, 1)
    logger.info(f"\n\nALL DONE at {end_time}. Total time: {mins_elapsed} minutes")


@error_wrapper()
def main(c: Config = None) -> None:
    """Main entrypoint function"""

    exit_if_already_running(c, Path(__file__).name)

    entrypoint()

    return None


if __name__ == "__main__":
    entrypoint()
